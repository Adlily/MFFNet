# -*- coding: utf-8 -*-
"""
Created on Sun Aug  2 22:20:26 2020

@author: Adlily
"""
import sys
sys.path.insert(1,'/ghome/suncl/code/package')
import os
import numpy as np
import openpyxl
import time
from config import config
split_format = config.SPLIT_FORMAT
def save_err_sample(err1,save_p):
    """
    保存错误分类的样本名称和类别
    """
    time1 = time.localtime(time.time())
    str1 = str(time1.tm_mon)+'_'+str(time1.tm_mday)+'_'+str(time1.tm_hour)+'_'+str(time1.tm_min)+'_'+str(time1.tm_sec)
    err_sample_file = os.path.join(save_p,"Error_simples_"+str1+'.xlsx')

    wb = openpyxl.workbook.Workbook()
    sheets=wb.worksheets
    sheet = sheets[0]
   
    #获取行数和列数
    rows = sheet.max_row
    cols = sheet.max_column
    sheet.cell(1,1).value = 'Simple_Class'
    sheet.cell(1,2).value = 'Simple_Name'
    sheet.cell(1,3).value = 'True_Label'
    sheet.cell(1,4).value = 'Predicted_Label'
    
    t_true  = err1[:,-2] #获取真是标签，并按真是标签进行排序
    index  = np.argsort(t_true)
    
    err = err1[index]
    
    for i in range(np.shape(err)[0]):
        neu_class = err[i,0].split(split_format)[0]
        sheet.cell(i+2,1).value = neu_class
        
        neu_name = err[i,0].split(split_format)[1]
        sheet.cell(i+2,2).value = neu_name
        
        for j in range(np.shape(err)[1]):
            if j!=0:
               sheet.cell(i+2,j+2).value = err[i,j]
    wb.save(err_sample_file)
    
def error_sample(pred_y,true_y,name):
    """
    返回被错误分类的样本名称
    """
    #index = np.arange(len(true_y))
    #err_index = index(pred_y != true_y)
    pred_y = np.array(pred_y)
    true_y = np.array(true_y)
    assert pred_y.shape==true_y.shape
    err_index = np.argwhere(pred_y!=true_y).flatten()
    err_sample = name[err_index]
    true_class = true_y[err_index]
    err_class = pred_y[err_index]
    err = np.vstack((err_sample,true_class))
    #print(err.shape)
    err_data = np.vstack((err,err_class)).T
    #print(err_data.shape)
    return err_data