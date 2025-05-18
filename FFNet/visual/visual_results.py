import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys
import numpy as np

from sklearn.metrics import roc_curve,auc,confusion_matrix,classification_report,roc_auc_score,precision_score,recall_score,precision_recall_fscore_support,precision_recall_curve
import itertools
from config import config
import time
classes = config.CLASS_NAME

def visual_all_results(test_true,test_pred,save_path,title=None,test_score=None,total=False):
    if not total:
        cm_path = os.path.join(save_path,'comfusion')
        re_path = os.path.join(save_path,'reports')
        all_saved_label_p = os.path.join(save_path,'saved_labels')
    else:
        cm_path,re_path,all_saved_label_p = save_path,save_path,save_path
    
    if not os.path.exists(cm_path):
        os.makedirs(cm_path)
    if not os.path.exists(re_path):
        os.makedirs(re_path)
    if not os.path.exists(all_saved_label_p):
        os.makedirs(all_saved_label_p)

    #metric.cm_plot(test_true,test_pred,cm_path)
    class_results(test_true,test_pred,re_path,y_score=test_score)


def write_data_excel(data,fileName,sheetname=None,startrow=None,truncate_sheet=False,**to_excel_kwargs):

    data_frame = pd.DataFrame(data)
    if os.path.exists(fileName):
        #print()
        with pd.ExcelWriter(fileName,mode='a',engine='openpyxl') as writer:
            #print(sheetname,writer.book.sheetnames)
            # 
            book = load_workbook(fileName)
            if sheetname is not None and sheetname in book.sheetnames:
                
                    writer.if_sheet_exists = 'replace'
                    df_old = pd.DataFrame(pd.read_excel(fileName,sheet_name=sheetname))
                    row_old = df_old.shape[0]

                    
                    writer.book = book
                    writer.sheets = dict((ws.title,ws) for ws in book.worksheets)

                    new_data = pd.concat((df_old,data_frame),axis=0)
                    new_data.to_excel(writer,sheet_name=sheetname,startrow=0,index=False,header=True)
                    #print('****************')
            else:
                    #print('------- ')
                    data_frame.to_excel(writer,sheet_name=sheetname,index=False)

    else:
        with pd.ExcelWriter(fileName) as write_pd:
            #print('&&&&&&&&&&&&&&&&')
            data_frame.to_excel(write_pd,sheet_name=sheetname)


# class Logger(object):
    # def __init__(self,fileN ="Default.log"):
        # self.terminal = sys.stdout
        # self.log = open(fileN,"a")
 
    # def write(self,message):
        # self.terminal.write(message)
        # self.log.write(message)
 
    # def flush(self):
        # pass

# def cm_plot_ratio(y, yp,save_p):
  # '''
  # plt the conmusion_contrix
  # '''
  # #y_true = T.as_tensor_variable(yp)
  # # labels_pred = model.predict_classes(data,verbose=0)

  # # Compute confusion matrix
  # cnf_matrix = confusion_matrix(y, yp)
  # #cnf_matrix = np.array([[254,99,2,17,128],[12,370,1,2,115],[12,28,12,0,43],[30,49,0,84,45],[55,118,10,1,315]])
  # #np.set_printoptions(precision=2)
  # cm_normalize = (cnf_matrix.astype('float') / cnf_matrix.sum(axis=1)[:, np.newaxis])
  # plt.figure()
  # plt.imshow(cm_normalize, interpolation='nearest', cmap=plt.cm.Greys)
  # plt.title('Confusion Matrix')
  # plt.colorbar()
  # tick_marks = np.arange(len(classes))
  # plt.xticks(tick_marks, classes, rotation=25)
  # plt.yticks(tick_marks, classes)
  # thresh = cnf_matrix.max() / 3

  # for i, j in itertools.product(range(cnf_matrix.shape[0]), range(cnf_matrix.shape[1])):
      # ratuo_ = cnf_matrix[i, j]/sum(cnf_matrix[i,:])*100
      # plt.text(j, i, format(ratuo_, ".2f"),
               # horizontalalignment="center",
               # color="white" if i==j else "black")

  # plt.ylabel('True Label')
  # plt.xlabel('Predicted Label')
  # #plt.tight_layout()

  # # plt.savefig(fig_dir + 'Normalized_confusion_matrix.svg')
  # time1 = time.localtime(time.time())
  # str1 = str(time1.tm_mon) + '_' + str(time1.tm_mday) + '_' + str(time1.tm_hour) + '_' + str(time1.tm_min) + '_' + str(
      # time1.tm_sec)
  # fig = plt.gcf()
  # fig.savefig(os.path.join(save_p,"{}_confusion_matrix_ratio.png".format(str1)))
  # plt.close()

# def cm_plot(y,yp,save_p):
    
    
    # #labels_pred = model.predict_classes(data,verbose=0)
    
    # # Compute confusion matrix
    # cnf_matrix = confusion_matrix(y,yp)
    # np.set_printoptions(precision=2)
    # cm_normalize = cnf_matrix.astype('float') / cnf_matrix.sum(axis=1)[:, np.newaxis]
    # plt.figure()
    # plt.imshow(cm_normalize, interpolation='nearest', cmap=plt.cm.Greys)
    # plt.title('Confusion Matrix')
    # plt.colorbar()
    # tick_marks = np.arange(len(classes))
    # plt.xticks(tick_marks, classes, rotation=25)
    # plt.yticks(tick_marks, classes)
    # thresh = cnf_matrix.max() / 2
    
    # for i, j in itertools.product(range(cnf_matrix.shape[0]), range(cnf_matrix.shape[1])):
        # plt.text(j, i, format(cnf_matrix[i, j], "d"),
                 # horizontalalignment="center",
                 # color="white" if  i==j  else "black")

    # plt.ylabel('True Label')
    # plt.xlabel('Predicted Label')
    # #plt.tight_layout()
    
    # #plt.savefig(fig_dir + 'Normalized_confusion_matrix.svg')
    # time1 = time.localtime(time.time())
    # str1 = str(time1.tm_mon)+'_'+str(time1.tm_mday)+'_'+str(time1.tm_hour)+'_'+str(time1.tm_min)+'_'+str(time1.tm_sec)
    # fig = plt.gcf()
    # fig.savefig(os.path.join(save_p,"{}_confusion_matrix.png".format(str1)))
    # plt.close()
    # cm_plot_ratio(y,yp,save_p)
    # #plt.show()

def compute_sen(Y_test,Y_pred):
    sen = []
    concat = confusion_matrix(Y_test,Y_pred)
    n = len(set(Y_test))
    for i in range(n):
        tp = concat[i][i]
        fn = np.sum(concat[i:]) - tp
        sen1 = tp /(tp+fn)
        sen.append(sen1)
    avg_sen = sum(sen)/n
    return sen, avg_sen


def compute_spe(Y_test,Y_pred):
    spe= []
    concat =  confusion_matrix(Y_test,Y_pred)
    n = len(set(Y_test))
    for i in range(n):
        number = np.sum(concat[:,:])
        tp = concat[i][i]
        fn = np.sum(concat[i,:]) - tp
        fp = np.sum(concat[:,i]) - tp
        tn =  number - tp -fn - fp
        spe1 = tn / (tn+fp)
        spe.append(spe1)
    avg_spe  = sum(spe)/n
    return spe,avg_spe

def class_results(y_true,y_preb,save_p,title=None,y_score=None):
    """
    show the classfication results for the per class
    avg_loss_acc:[avg_tra_loss,acg_tra_acc,avg_test_loss,avg_test_acc]
    """
    time1 = time.localtime(time.time())
    str1 = str(time1.tm_mon)+'_'+str(time1.tm_mday)+'_'+str(time1.tm_hour)+'_'+str(time1.tm_min)+'_'+str(time1.tm_sec)
    results = classification_report(y_true, y_preb,digits=6)
    
    if title is not None:
        f_name = os.path.join(save_p,"{}_class_result_{}.txt".format(str1,title))
    else:
        f_name = os.path.join(save_p,"{}_class_result_.txt".format(str1))

    with open(f_name,"w") as f:
        f.write(results+"\n")
        #write confusion matrix to file
        cnf_matrix = np.array(confusion_matrix(y_true, y_preb))
        cm_normalize = np.array(cnf_matrix.astype('float') / cnf_matrix.sum(axis=1)[:, np.newaxis]*100)

        f.write('confusiom matrix ratio is:\n')
        np.savetxt(f, np.reshape(np.array(classes), (1, len(classes))), fmt='%s', delimiter='\t')
        f.write('\n')

        np.savetxt(f, cm_normalize, fmt='%.2f', delimiter='\t')
        f.write('\n')
        f.write('confusiom matrix is:\n')
        f.write('confusiom matrix ratio is:\n')
        np.savetxt(f, cnf_matrix, fmt='%d', delimiter='\t')
        f.write('\n')
        
        #计算敏感度和特异性
        senti,avg_sen = compute_sen(y_true,y_preb)
        specity, avg_spec = compute_spe(y_true,y_preb)

        f.write('avg sensitity is:{}, sensitity for per class is \n'.format(avg_sen))
        np.savetxt(f,senti,fmt='%.6f',delimiter='   ')
        f.write('\n')
        f.write('avg spectity is:{}, spectity for per class is \n'.format(avg_spec))
        np.savetxt(f,specity,fmt='%.6f',delimiter='   ')
        f.write('\n')

        #计算ROC
        if y_score is not None:
            fpr,tpr,thresholds =  roc_curve(y_true,y_score,pos_label=2)
            roc = auc(fpr,tpr)
            f.write('auc is {}\n'.format(roc))

            f.write('fpr is \n')
            np.savetxt(f,np.array(fpr),fmt='%.6f',delimiter='   ')
            f.write('\n')
            f.write('tpr is \n')
            np.savetxt(f,np.array(tpr),fmt='%.6f',delimiter='   ')
            f.write('\n')
            f.write('threshold is \n')
            np.savetxt(f,np.array(thresholds),fmt='%.6f',delimiter='   ')
            f.write('\n')
            