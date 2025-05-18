#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/9/7 11:42
# @Author  : Sun Chunli
# @Software: PyCharm
import os
import time
import openpyxl
import numpy as np
from config import config
import sys
sys.path.insert(1,'/ghome/suncl/code/package')

image2 = config.IMAGE2
from config import config
split_format = config.SPLIT_FORMAT

def write_results(result,fig_dir1,cv_num):

    with open(os.path.join(fig_dir1, 'multi_input' + '_' + str(cv_num) + '.txt'), 'w') as f:
        acc_cv_num = result.history['acc']
        loss_cv_num = result.history['loss']
        val_acc_cv_num = result.history['val_acc']
        val_loss_cv_num = result.history['val_loss']
        f.write('Train_acc\tTrain_loss\tVal_acc\tVal_loss\n')
        for i in range(len(acc_cv_num)):
            f.write(str(acc_cv_num[i]))
            f.write('\t')
            f.write(str(loss_cv_num[i]))
            f.write('\t')
            f.write(str(val_acc_cv_num[i]))
            f.write('\t')
            f.write(str(val_loss_cv_num[i]))
            f.write('\n')

def write_loss_acc(data,fig_dir1,cv_num):
    with open(os.path.join(fig_dir1, 'val_loss_acc' + '.txt'), 'a') as f:
        if cv_num == 0:
            f.write('CV_num\tAcc\tLoss\n')
        f.write(str(cv_num))
        f.write('\t')
        for i in range(len(data)):


           f.write(str(data[i][:6]))
           f.write('\t')
        f.write('\n')

def write_all_acc(data,filename):
    with open(filename, 'a') as f:
        f.write('Repeat_Time')
        f.write('\tClf_acc\n')
        for i in range(len(data)):
           f.write(str(i))
           f.write('\t')
           f.write(str(data[i]))
           f.write('\n')

def save_feture(save_p,err,label,name):

    """
     保存错误分类的样本名称和类别
     """
    
    time1 = time.localtime(time.time())
    str1 = str(time1.tm_mon) + '_' + str(time1.tm_mday) + '_' + str(time1.tm_hour) + '_' + str(
        time1.tm_min) + '_' + str(time1.tm_sec)
    #err_sample_file = os.path.join(save_p,  str1 + '_feat.xlsx')
    
    if os.path.exists(save_p):
        print('file exists')
        wb = openpyxl.load_workbook(save_p)
    else:
        print('file not exists')
        wb = openpyxl.workbook.Workbook()
    sheets = wb.worksheets
    sheet = sheets[0]

    # 获取行数和列数
    rows = sheet.max_row
    print(rows)
    cols = sheet.max_column
    
    #label = np.argmax(label,axis=1)
    
    for i in range(np.shape(err)[0]):
        sheet.cell(rows + i+1 , 1 ).value = name[i].split(split_format)[-1]+'.CNG.swc'
        for j in range(np.shape(err)[1]):
            sheet.cell(rows+i+1 , j+2).value = err[i][j]
        j=j+1
        sheet.cell(rows + i+1,  j+2 ).value = str(label[i])
        
        
    wb.save(save_p)
    save_p_p = save_p.split('.')[0]
    wb.save(save_p_p+str1+'.xlsx')
    
    
if __name__=="__main__":

    path = '/ghome/suncl/code/projected_image/text1.xlsx'
    save_feture(path,[[1,2,3],[4,5,6]],[0,1],['1','3'])